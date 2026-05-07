import openpyxl
from datetime import datetime
import os

class ExcelFiller:
    def __init__(self, template_path):
        self.template_path = template_path

    def fill_bill_data(self, data, output_path):
        """
        Fills extracted data into the Excel template with professional formatting.
        """
        wb = openpyxl.load_workbook(self.template_path)
        ws = wb.active

        # 1. Update basic info (Meter 1)
        if data.get("consumer_name"):
            ws["D1"] = data["consumer_name"]
        
        if data.get("consumer_number"):
            # Format as number to avoid scientific notation (E+11)
            try:
                val = float(str(data["consumer_number"]).replace(" ", ""))
                ws["D2"] = val
                ws["D2"].number_format = '0' # Force whole number format
            except:
                ws["D2"] = data["consumer_number"]

        if data.get("fixed_charges"):
            try:
                ws["D3"] = float(data["fixed_charges"])
            except:
                ws["D3"] = data["fixed_charges"]

        if data.get("sanctioned_load"):
            ws["D4"] = data["sanctioned_load"]
            
        if data.get("connection_type"):
            ws["D5"] = data["connection_type"]

        # 2. Find the row for the billing month
        billing_month_str = data.get("billing_month")
        if billing_month_str:
            target_row = self._find_month_row(ws, billing_month_str)
            if target_row:
                if data.get("units_consumed") is not None:
                    try:
                        ws[f"D{target_row}"] = float(data["units_consumed"])
                    except:
                        ws[f"D{target_row}"] = data["units_consumed"]
                
                if data.get("bill_amount") is not None:
                    try:
                        ws[f"E{target_row}"] = float(data["bill_amount"])
                    except:
                        ws[f"E{target_row}"] = data["bill_amount"]

        # 3. FIX: Adjust column widths to prevent #####
        # Widen columns C, G (Months) and D, E, H, I (Data)
        for col_letter in ['C', 'D', 'E', 'G', 'H', 'I']:
            ws.column_dimensions[col_letter].width = 18

        # Save the file
        wb.save(output_path)
        return output_path

    def _find_month_row(self, ws, month_str):
        try:
            month_str = month_str.replace("-", " ").replace("/", " ")
            dt = None
            for fmt in ("%B %Y", "%b %Y", "%m %Y"):
                try:
                    dt = datetime.strptime(month_str, fmt)
                    break
                except ValueError:
                    continue
            
            if not dt:
                return None

            for row in range(9, 25):
                cell_val = ws[f"C{row}"].value
                if isinstance(cell_val, datetime):
                    if cell_val.year == dt.year and cell_val.month == dt.month:
                        return row
                elif isinstance(cell_val, str):
                    try:
                        # Common Excel date string format
                        c_dt = datetime.strptime(cell_val.split(" ")[0], "%Y-%m-%d")
                        if c_dt.year == dt.year and c_dt.month == dt.month:
                            return row
                    except:
                        pass
        except Exception as e:
            print(f"Error in _find_month_row: {e}")
        
        return None
